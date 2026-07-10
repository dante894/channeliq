from datetime import datetime, timedelta
import requests

YOUTUBE_TOKEN_URL = "https://oauth2.googleapis.com/token"
YOUTUBE_SCOPES = [
    "https://www.googleapis.com/auth/youtube.readonly",
    "https://www.googleapis.com/auth/yt-analytics.readonly",
]


class YouTubeAPIError(Exception):
    def __init__(self, message, needs_reconnect=False):
        super().__init__(message)
        self.needs_reconnect = needs_reconnect


def get_auth_url(client_id, redirect_uri):
    import urllib.parse
    params = {
        "client_id": client_id,
        "redirect_uri": redirect_uri,
        "response_type": "code",
        "scope": " ".join(YOUTUBE_SCOPES),
        "access_type": "offline",
        "prompt": "consent",
    }
    return "https://accounts.google.com/o/oauth2/v2/auth?" + urllib.parse.urlencode(params)


def exchange_code(code, client_id, client_secret, redirect_uri):
    resp = requests.post(YOUTUBE_TOKEN_URL, data={
        "code": code, "client_id": client_id,
        "client_secret": client_secret,
        "redirect_uri": redirect_uri,
        "grant_type": "authorization_code",
    })
    data = resp.json()
    if "error" in data:
        raise YouTubeAPIError(data.get("error_description", data["error"]))
    return data


def refresh_token(refresh_tok, client_id, client_secret):
    resp = requests.post(YOUTUBE_TOKEN_URL, data={
        "refresh_token": refresh_tok, "client_id": client_id,
        "client_secret": client_secret, "grant_type": "refresh_token",
    })
    data = resp.json()
    if "error" in data:
        message = data.get("error_description", data["error"])
        needs_reconnect = data.get("error") == "invalid_grant"
        raise YouTubeAPIError(message, needs_reconnect=needs_reconnect)
    return data


def get_valid_token(channel, client_id, client_secret):
    from app.extensions import db
    if channel.is_token_expired():
        data = refresh_token(channel.refresh_token, client_id, client_secret)
        channel.access_token = data["access_token"]
        channel.token_expiry = datetime.utcnow() + timedelta(seconds=data.get("expires_in", 3600))
        db.session.commit()
    return channel.access_token


def fetch_channel_info(access_token):
    resp = requests.get(
        "https://www.googleapis.com/youtube/v3/channels",
        params={"part": "snippet,statistics", "mine": "true"},
        headers={"Authorization": f"Bearer {access_token}"},
    )
    data = resp.json()
    if "error" in data:
        raise YouTubeAPIError(data["error"].get("message", "Error de API"))
    items = data.get("items", [])
    if not items:
        raise YouTubeAPIError("No se encontró ningún canal.")
    item = items[0]
    stats = item["statistics"]
    return {
        "channel_id": item["id"],
        "channel_name": item["snippet"]["title"],
        "channel_thumbnail": item["snippet"]["thumbnails"].get("default", {}).get("url", ""),
        "subscriber_count": int(stats.get("subscriberCount", 0)),
        "video_count": int(stats.get("videoCount", 0)),
        "view_count": int(stats.get("viewCount", 0)),
    }


def fetch_analytics(access_token, channel_id, days=7):
    end = datetime.utcnow().date()
    start = end - timedelta(days=days)
    resp = requests.get(
        "https://youtubeanalytics.googleapis.com/v2/reports",
        params={
            "ids": f"channel=={channel_id}",
            "startDate": start.isoformat(),
            "endDate": end.isoformat(),
            "metrics": "views,estimatedMinutesWatched,averageViewDuration,subscribersGained,subscribersLost",
            "dimensions": "day",
            "sort": "day",
        },
        headers={"Authorization": f"Bearer {access_token}"},
    )
    data = resp.json()
    if "error" in data:
        raise YouTubeAPIError(data["error"].get("message", "Error Analytics API"))

    headers = [h["name"] for h in data.get("columnHeaders", [])]
    rows = data.get("rows", [])
    daily, totals = [], {"views": 0, "minutes_watched": 0, "subs_gained": 0, "subs_lost": 0}

    for row in rows:
        e = dict(zip(headers, row))
        daily.append({
            "date": e.get("day"),
            "views": int(e.get("views", 0)),
            "minutes_watched": int(e.get("estimatedMinutesWatched", 0)),
            "avg_duration": int(e.get("averageViewDuration", 0)),
            "subs_gained": int(e.get("subscribersGained", 0)),
            "subs_lost": int(e.get("subscribersLost", 0)),
        })
        totals["views"] += int(e.get("views", 0))
        totals["minutes_watched"] += int(e.get("estimatedMinutesWatched", 0))
        totals["subs_gained"] += int(e.get("subscribersGained", 0))
        totals["subs_lost"] += int(e.get("subscribersLost", 0))

    return {"daily": daily, "totals": totals, "days": days}


def fetch_videos(access_token, channel_id, max_results=10):
    search = requests.get(
        "https://www.googleapis.com/youtube/v3/search",
        params={"part": "snippet", "channelId": channel_id, "type": "video",
                "order": "date", "maxResults": max_results},
        headers={"Authorization": f"Bearer {access_token}"},
    ).json()
    if "error" in search:
        return []
    ids = [i["id"]["videoId"] for i in search.get("items", [])]
    if not ids:
        return []
    vdata = requests.get(
        "https://www.googleapis.com/youtube/v3/videos",
        params={"part": "snippet,statistics", "id": ",".join(ids)},
        headers={"Authorization": f"Bearer {access_token}"},
    ).json()
    result = []
    for item in vdata.get("items", []):
        s = item.get("statistics", {})
        result.append({
            "id": item["id"],
            "title": item["snippet"]["title"],
            "published_at": item["snippet"]["publishedAt"][:10],
            "thumbnail": item["snippet"]["thumbnails"].get("medium", {}).get("url", ""),
            "views": int(s.get("viewCount", 0)),
            "likes": int(s.get("likeCount", 0)),
            "comments": int(s.get("commentCount", 0)),
        })
    return result


def build_excel(channel_info, analytics, videos):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
    from openpyxl.chart import LineChart, Reference
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.utils import get_column_letter

    # ── Paleta (alineada al branding de ChannelIQ) ───────────────────────────
    NAVY = "1A2235"
    NAVY_DARK = "0A0E17"
    ACCENT = "4F8EF7"
    ACCENT_LIGHT = "EAF1FE"
    GREEN = "16A34A"
    RED = "DC2626"
    GRAY_TEXT = "6B7FA3"
    LINE = "D8E1EF"
    ROW_ALT = "F5F8FD"

    FONT = "Arial"
    thin = Side(style="thin", color=LINE)
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    def f(size=10, bold=False, color="1A2235", italic=False):
        return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)

    def style_header_row(ws, row, first_col, last_col, height=22):
        ws.row_dimensions[row].height = height
        for col in range(first_col, last_col + 1):
            c = ws.cell(row, col)
            c.font = f(10, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(vertical="center", horizontal="left" if col == first_col else "right")
            c.border = border_all

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════════════
    # Hoja 1: Resumen
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.sheet_properties.tabColor = ACCENT
    ws1.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGH", [2, 20, 20, 20, 20, 20, 20, 2]):
        ws1.column_dimensions[col].width = w

    # Banner de título
    ws1.merge_cells("B2:G3")
    title_cell = ws1["B2"]
    title_cell.value = "ChannelIQ — Reporte de Canal"
    title_cell.font = f(20, bold=True, color="FFFFFF")
    title_cell.fill = GradientFill(stop=(NAVY_DARK, NAVY), degree=90)
    title_cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    for row in (2, 3):
        for col in range(2, 8):
            cell = ws1.cell(row, col)
            if cell.value is None:
                cell.fill = GradientFill(stop=(NAVY_DARK, NAVY), degree=90)
    ws1.row_dimensions[2].height = 20
    ws1.row_dimensions[3].height = 20

    ws1.merge_cells("B4:G4")
    sub_cell = ws1["B4"]
    sub_cell.value = (f"{channel_info['channel_name']}  ·  "
                      f"Generado el {datetime.utcnow().strftime('%d/%m/%Y a las %H:%M')} UTC")
    sub_cell.font = f(10, italic=True, color=GRAY_TEXT)
    ws1.row_dimensions[4].height = 18

    # Tarjetas KPI
    kpis = [
        ("SUSCRIPTORES", channel_info["subscriber_count"]),
        ("VIDEOS TOTALES", channel_info["video_count"]),
        ("VISTAS HISTÓRICAS", channel_info["view_count"]),
        (f"VISTAS ({analytics['days']}D)", analytics["totals"]["views"]),
    ]
    kpi_row_label, kpi_row_value = 6, 7
    kpi_cols = [("B", "C"), ("D", "E"), ("F", "G")]
    # Usamos 4 tarjetas repartidas en B..I (ampliamos si hace falta)
    ws1.column_dimensions["I"].width = 20
    kpi_cols = [("B", "C"), ("D", "E"), ("F", "G"), ("H", "I")]
    for (start, end), (label, value) in zip(kpi_cols, kpis):
        ws1.merge_cells(f"{start}{kpi_row_label}:{end}{kpi_row_label}")
        ws1.merge_cells(f"{start}{kpi_row_value}:{end}{kpi_row_value}")
        lc = ws1[f"{start}{kpi_row_label}"]
        vc = ws1[f"{start}{kpi_row_value}"]
        lc.value = label
        lc.font = f(8.5, bold=True, color=GRAY_TEXT)
        lc.alignment = Alignment(horizontal="center")
        vc.value = value
        vc.number_format = "#,##0"
        vc.font = f(22, bold=True, color=ACCENT)
        vc.alignment = Alignment(horizontal="center")
        for r in (kpi_row_label, kpi_row_value):
            for col in (start, end):
                ws1[f"{col}{r}"].fill = PatternFill("solid", fgColor=ACCENT_LIGHT)
        # borde exterior de la tarjeta
        for col in (start, end):
            ws1[f"{col}{kpi_row_label}"].border = Border(top=Side("thin", color=ACCENT), left=Side("thin", color=ACCENT) if col == start else None, right=Side("thin", color=ACCENT) if col == end else None)
            ws1[f"{col}{kpi_row_value}"].border = Border(bottom=Side("thin", color=ACCENT), left=Side("thin", color=ACCENT) if col == start else None, right=Side("thin", color=ACCENT) if col == end else None)
    ws1.row_dimensions[kpi_row_label].height = 18
    ws1.row_dimensions[kpi_row_value].height = 34

    # Tabla "Últimos N días"
    period_row = 9
    ws1.merge_cells(f"B{period_row}:C{period_row}")
    hcell = ws1[f"B{period_row}"]
    hcell.value = f"ÚLTIMOS {analytics['days']} DÍAS"
    hcell.font = f(10, bold=True, color="FFFFFF")
    hcell.fill = PatternFill("solid", fgColor=NAVY)
    hcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[period_row].height = 20
    for col in ("D", "E", "F", "G", "H", "I"):
        ws1[f"{col}{period_row}"].fill = PatternFill("solid", fgColor=NAVY)

    t = analytics["totals"]
    net = t["subs_gained"] - t["subs_lost"]
    rows_data = [
        ("Vistas", t["views"], None),
        ("Minutos vistos", t["minutes_watched"], None),
        ("Subs. ganados", t["subs_gained"], GREEN),
        ("Subs. perdidos", t["subs_lost"], RED),
        ("Subs. netos", None, GREEN if net >= 0 else RED),  # fórmula abajo
    ]
    r0 = period_row + 1
    for i, (label, value, color) in enumerate(rows_data):
        r = r0 + i
        lc, vc = ws1[f"B{r}"], ws1[f"C{r}"]
        lc.value = label
        lc.font = f(10, color="1A2235")
        vc.font = f(10, bold=True, color=color or "1A2235")
        vc.number_format = "+#,##0;-#,##0;0"
        fill = PatternFill("solid", fgColor=ROW_ALT) if i % 2 == 0 else PatternFill(fill_type=None)
        for col in ("B", "C"):
            ws1[f"{col}{r}"].fill = fill
            ws1[f"{col}{r}"].border = border_all
        if label == "Subs. netos":
            vc.value = f"=C{r0 + 2}-C{r0 + 3}"  # ganados - perdidos
        else:
            vc.value = value

    ws1.freeze_panes = "B6"

    # ══════════════════════════════════════════════════════════════════════
    # Hoja 2: Analytics diarios
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Analytics diarios")
    ws2.sheet_properties.tabColor = GREEN
    ws2.sheet_view.showGridLines = False
    cols = ["Fecha", "Vistas", "Minutos vistos", "Duración prom (seg)", "Subs. ganados", "Subs. perdidos"]
    widths = [14, 12, 16, 20, 14, 14]
    for ci, (h, w) in enumerate(zip(cols, widths), 1):
        letter = get_column_letter(ci)
        ws2.column_dimensions[letter].width = w
        ws2.cell(1, ci, h)
    style_header_row(ws2, 1, 1, len(cols))

    for ri, d in enumerate(analytics["daily"], 2):
        vals = [d["date"], d["views"], d["minutes_watched"], d["avg_duration"], d["subs_gained"], d["subs_lost"]]
        fill = PatternFill("solid", fgColor=ROW_ALT) if ri % 2 == 0 else PatternFill(fill_type=None)
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(ri, ci, v)
            c.font = f(9.5)
            c.border = border_all
            c.fill = fill
            if ci > 1:
                c.number_format = "#,##0"
                c.alignment = Alignment(horizontal="right")

    ws2.freeze_panes = "A2"
    last_row = len(analytics["daily"]) + 1
    if last_row > 1:
        ws2.auto_filter.ref = f"A1:F{last_row}"

    if len(analytics["daily"]) > 1:
        chart = LineChart()
        chart.title = "Vistas diarias"
        chart.style = 10
        chart.height = 9
        chart.width = 22
        chart.y_axis.title = "Vistas"
        chart.x_axis.title = "Fecha"
        data_ref = Reference(ws2, min_col=2, min_row=1, max_row=last_row)
        cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=last_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        series = chart.series[0]
        series.graphicalProperties.line.width = 22000  # ~2.2pt
        series.graphicalProperties.line.solidFill = ACCENT
        series.marker.symbol = "circle"
        series.marker.size = 5
        series.marker.graphicalProperties.solidFill = ACCENT
        series.marker.graphicalProperties.line.solidFill = ACCENT
        series.smooth = True
        chart.legend = None
        ws2.add_chart(chart, f"H2")

    # ══════════════════════════════════════════════════════════════════════
    # Hoja 3: Videos
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Videos")
    ws3.sheet_properties.tabColor = "FBBF24"
    ws3.sheet_view.showGridLines = False
    vh = ["Título", "Fecha", "Vistas", "Likes", "Comentarios"]
    vw = [52, 14, 14, 12, 14]
    for ci, (h, w) in enumerate(zip(vh, vw), 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w
        ws3.cell(1, ci, h)
    style_header_row(ws3, 1, 1, len(vh))

    for ri, v in enumerate(videos, 2):
        vals = [v["title"], v["published_at"], v["views"], v["likes"], v["comments"]]
        fill = PatternFill("solid", fgColor=ROW_ALT) if ri % 2 == 0 else PatternFill(fill_type=None)
        for ci, val in enumerate(vals, 1):
            c = ws3.cell(ri, ci, val)
            c.font = f(9.5)
            c.border = border_all
            c.fill = fill
            if ci >= 3:
                c.number_format = "#,##0"
                c.alignment = Alignment(horizontal="right")
            elif ci == 1:
                c.alignment = Alignment(horizontal="left")

    ws3.freeze_panes = "A2"
    last_video_row = len(videos) + 1
    if last_video_row > 1:
        ws3.auto_filter.ref = f"A1:E{last_video_row}"
        # Mini barras de datos sobre "Vistas" para comparar de un vistazo
        rule = DataBarRule(start_type="min", end_type="max", color=ACCENT.replace("4F8EF7", "4F8EF7"))
        ws3.conditional_formatting.add(f"C2:C{last_video_row}", rule)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
